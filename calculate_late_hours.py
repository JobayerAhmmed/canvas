from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

import config
from config import columns
import utils



def update_late_hours():
    print('Calculating late hours ...')

    date_format = '%Y-%m-%d %H:%M:%S'
    extended_due_date = datetime.strptime(config.extended_due_date_str, date_format)

    df = pd.read_excel(config.file_grades)

    df[columns.late_hours] = df.apply(
        lambda row: calculate_late_hours(
            extended_due_date, row[columns.submitted_at]),
        axis=1)
        
    late_hours_dict = dict(zip(df[columns.student_id], df[columns.late_hours]))

    wb = load_workbook(config.file_grades)
    ws = wb.active 
    
    student_id_col = utils.find_column(ws, columns.student_id)
    late_hours_col = utils.find_column(ws, columns.late_hours)

    if student_id_col and late_hours_col:
        for row in range(2, ws.max_row + 1):
            student_id = ws.cell(row=row, column=student_id_col).value
            if student_id in late_hours_dict:
                ws.cell(row=row, column=late_hours_col,
                        value=late_hours_dict[student_id])

    wb.save(config.file_grades)


def calculate_late_hours(extended_due_date, submitted_date):
    if submitted_date > extended_due_date:
        late_hours = (submitted_date - extended_due_date).total_seconds() / 3600
    else:
        late_hours = ''
    
    return late_hours


if __name__ == "__main__":
    update_late_hours()
