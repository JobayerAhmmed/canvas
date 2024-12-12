import os
import shutil
import time

from openpyxl import load_workbook
import progressbar

import config
from config import columns
import utils



def create_output_txt():
    print('Creating student feedback files ...')

    wb = load_workbook(config.file_grades_unsorted, data_only=True)
    ws = wb.active

    student_name_col = utils.find_column(ws, columns.student_name)
    submitted_at_col = utils.find_column(ws, columns.submitted_at)
    late_deduct_col = utils.find_column(ws, columns.late_deduct)
    early_points_col = utils.find_column(ws, columns.early_points)
    auto_col = utils.find_column(ws, columns.auto)
    manual1_col = utils.find_column(ws, columns.manual1)
    manual2_col = utils.find_column(ws, columns.manual2)
    manual3_col = utils.find_column(ws, columns.manual3)
    manual4_col = utils.find_column(ws, columns.manual4)
    total_col = utils.find_column(ws, columns.total)
    comments_col = utils.find_column(ws, columns.comments)    

    for row in range(2, ws.max_row + 1):
        student_name = ws.cell(row=row, column=student_name_col).value
        submitted_at = ws.cell(row=row, column=submitted_at_col).value
        late_deduct = ws.cell(row=row, column=late_deduct_col).value
        early_points = ws.cell(row=row, column=early_points_col).value
        auto = ws.cell(row=row, column=auto_col).value
        manual1 = ws.cell(row=row, column=manual1_col).value
        manual2 = ws.cell(row=row, column=manual2_col).value
        manual3 = ws.cell(row=row, column=manual3_col).value
        manual4 = ws.cell(row=row, column=manual4_col).value
        total = ws.cell(row=row, column=total_col).value
        comments = ws.cell(row=row, column=comments_col).value

        output_file_name = f'{student_name}_HW{config.hw_no}.txt'
        if submitted_at:
            output_content = config.output_head + '\n\n'
            log_file = os.path.join(config.submissions_with_grader_dir,
                                    student_name, 'log.txt')
            if os.path.isfile(log_file):
                output_content += open(log_file).read().strip() + '\n'

            output_content += '-----------------------------------------------\n\n\n'
            output_content += 'GRADES\n---------------------------------------\n'
            output_content += f'Early submission points (10%): {early_points}\n'
            output_content += f'Late submission points (10%): {late_deduct}\n'
            output_content += f'Automated score: {auto} / 70\n'

            manual_scores = manual1 + manual2 + manual3 + manual4
            output_content += f'Manual checks: {manual_scores} / 30\n'
            output_content += 'Comments:\n-----------\n'
            output_content += f'{comments}\n\n\n-----------------------------------------\n'
            output_content += f'TOTAL GRADE: {total}\n'

            output_file = os.path.join(config.submissions_with_grader_dir,
                                       student_name, output_file_name)
            with open(output_file, 'w') as f:
                f.write(output_content)


def copy_output_txt():
    print('Copying student feedback files ...')

    os.makedirs(config.student_reports_dir, exist_ok=True)

    wb = load_workbook(config.file_grades_unsorted, data_only=True)
    ws = wb.active

    student_name_col = utils.find_column(ws, columns.student_name)
    submitted_at_col = utils.find_column(ws, columns.submitted_at)

    bar = progressbar.ProgressBar(max_value=ws.max_row,
                                  redirect_stdout=True).start()
    i = 1
    for row in range(2, ws.max_row + 1):
        student_name = ws.cell(row=row, column=student_name_col).value
        submitted_at = ws.cell(row=row, column=submitted_at_col).value
        if not submitted_at:
            continue
        
        file_name = f'{student_name}_HW{config.hw_no}.txt'
        src_file = os.path.join(config.submissions_with_grader_dir,
                                student_name, file_name)
        dst_file = os.path.join(config.student_reports_dir, file_name)
        try:
            shutil.copy(src_file, dst_file)
        except FileNotFoundError as fnfe:
            print(f"Error: The file {src_file} does not exist: {fnfe}")
        except PermissionError as pe:
            print(f"Error: Permission denied to access {dst_file}: {pe}")
        except Exception as e:
            print(f"An error occurred: {e}")
        time.sleep(1)
        i += 1
        bar.update(i)
    bar.finish()


if __name__ == "__main__":
    create_output_txt()
    copy_output_txt()
