from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

import config
from config import columns
import utils



def update_early_days():
    print('Calculating early days ...')

    date_format = '%Y-%m-%d %H:%M:%S'
    due_date = datetime.strptime(config.due_date_str, date_format)

    df = pd.read_excel(config.file_grades)

    df[columns.early_days] = df.apply(
        lambda row: calculate_early_days(
            due_date, row[columns.submitted_at]),
        axis=1)
        
    early_days_dict = dict(zip(df[columns.student_id], df[columns.early_days]))

    wb = load_workbook(config.file_grades)
    ws = wb.active

    student_id_col = utils.find_column(ws, columns.student_id)
    early_days_col = utils.find_column(ws, columns.early_days)
    
    if student_id_col and early_days_col:
        for row in range(2, ws.max_row + 1):
            student_id = ws.cell(row=row, column=student_id_col).value
            if student_id in early_days_dict:
                ws.cell(row=row, column=early_days_col,
                        value=early_days_dict[student_id])

    wb.save(config.file_grades)


def calculate_early_days(due_date, submitted_date):
    time_difference = due_date - submitted_date
    if time_difference.days > 0:
        early_days = time_difference.days
    else:
        early_days = ''

    return early_days


if __name__ == "__main__":
    update_early_days()
