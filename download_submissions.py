import os
import requests

import pandas as pd
from canvasapi import Canvas
from openpyxl import load_workbook
import progressbar

import config
from config import columns
import utils



def download_submissions():
    """Download students submissions and update Submitted At column."""
    print('Downloading students submissions ...')

    canvas = Canvas(config.api_url, config.api_key)
    course = canvas.get_course(config.course_id)
    assignment = course.get_assignment(config.assignment_id)

    os.makedirs(config.submissions_dir, exist_ok=True)

    df = pd.read_excel(config.file_grades)

    # Convert 'submitted_at' column to datetime type
    df[columns.submitted_at] = pd.to_datetime(df[columns.submitted_at],
                                              errors='coerce')

    bar = progressbar.ProgressBar(max_value=df.shape[0],
                                  redirect_stdout=True).start()
    for index, row in df.iterrows():
        student_name = row[columns.student_name]
        student_id = row[columns.student_id]

        submission = assignment.get_submission(student_id)

        if submission and submission.attachments:
            submitted_at = submission.submitted_at # UTC format: 2024-09-22T04:19:51Z
            submitted_at_local = utils.convert_utc_to_iowa(submitted_at)
            df.at[index, columns.submitted_at] = pd.to_datetime(submitted_at_local)
            latest_attachment = submission.attachments[-1]
            file_url = latest_attachment.url
            file_extension = os.path.splitext(latest_attachment.filename)[1]
            file_name = f"{student_name}{file_extension}"
            file_path = os.path.join(config.submissions_dir, file_name)

            response = requests.get(file_url, headers={'Authorization': f'Bearer {config.api_key}'})
            with open(file_path, 'wb') as file:
                file.write(response.content)
        else:
            print(f'No submission found for {student_name}')
        bar.update(index)
    bar.finish()

    submitted_at_dict = dict(zip(df[columns.student_id], df[columns.submitted_at]))

    wb = load_workbook(config.file_grades)
    ws = wb.active

    student_id_col = utils.find_column(ws, columns.student_id)
    submitted_at_col = utils.find_column(ws, columns.submitted_at)

    if student_id_col and submitted_at_col:
        for row in range(2, ws.max_row + 1):
            student_id = ws.cell(row=row, column=student_id_col).value
            if student_id in submitted_at_dict:
                ws.cell(row=row, column=submitted_at_col,
                        value=submitted_at_dict[student_id])

    wb.save(config.file_grades)


if __name__ == "__main__":
    download_submissions()
