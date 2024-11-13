
from openpyxl import load_workbook, Workbook

import config



def unsort_grades():
    print('Sorting students based on original data ...')

    wb_students = load_workbook(config.file_students)
    wb_grades = load_workbook(config.file_grades, data_only=True)

    ws_students = wb_students.active
    ws_grades = wb_grades.active

    student_names = [cell.value for cell in ws_students['A'] if cell.value is not None]

    data_to_sort = []
    header = [cell.value for cell in ws_grades[1]]
    data_to_sort.append(header)

    student_data_mapping = {row[0]: row for row in ws_grades.iter_rows(min_row=2, values_only=True)}

    for name in student_names:
        if name in student_data_mapping:
            data_to_sort.append(student_data_mapping[name])

    wb_sorted = Workbook()
    ws_sorted = wb_sorted.active

    for row in data_to_sort:
        ws_sorted.append(row)

    wb_sorted.save(config.file_grades_unsorted)


if __name__ == "__main__":
    unsort_grades()
