import os
import shutil
from datetime import datetime

from openpyxl import Workbook, load_workbook

import config
from config import columns



def create_excel_with_student_names_and_headers():
    print('Creating grades.xlsx ...')

    if os.path.isfile(config.file_grades):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        dir_path = os.path.dirname(config.file_grades)
        filename = os.path.basename(config.file_grades) + '_' + timestamp + '.xlsx'
        backup_file_path = os.path.join(dir_path, filename)
        shutil.copy(config.file_grades, backup_file_path)

    wb_students = load_workbook(filename=config.file_students)
    ws_students = wb_students.active

    wb_grades = Workbook()
    ws_grades = wb_grades.active

    for index, header in enumerate(columns, start=1):
        ws_grades.cell(row=1, column=index, value=header)

    for row_index, row in enumerate(ws_students.iter_rows(min_row=2, max_col=2,
                                                          values_only=True),
                                                          start=2):
        ws_grades.cell(row=row_index, column=1, value=row[0])
        ws_grades.cell(row=row_index, column=2, value=row[1])

    wb_grades.save(filename=config.file_grades)


def sort_by_student_name():
    workbook = load_workbook(filename=config.file_grades)
    worksheet = workbook.active

    header = [cell.value for cell in worksheet[1]]
    data_rows = [
        [cell for cell in row]
        for row in worksheet.iter_rows(min_row=2, values_only=True)
    ]

    try:
        student_name_col = header.index(columns[0]) # Student Name
    except ValueError:
        print(f"Column '{columns[0]}' not found!")
        workbook.close()
        exit()

    data_rows.sort(key=lambda x: x[student_name_col])

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row,
                                   max_col=worksheet.max_column):
        for cell in row:
            cell.value = None

    for row_index, row_data in enumerate(data_rows, start=2):
        for col_index, value in enumerate(row_data, start=1):
            worksheet.cell(row=row_index, column=col_index, value=value)

    workbook.save(filename=config.file_grades)



if __name__ == "__main__":
    create_excel_with_student_names_and_headers()
    sort_by_student_name()
